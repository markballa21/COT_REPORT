import pandas as pd
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os


class Excel():

    def __init__(self, name, dataFrame):
        self.name = name
        self.dataFrame = dataFrame
        self.write_excel()

    def write_excel(self):
        #the path of the file
        path = 'COT_REPORT.xlsx'
        #check if path exists, we use mode 'a' to append and replace sheets that exist
        #else we write a new excel file
        if os.path.isfile(path):
            writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='replace')
        else:
            writer = pd.ExcelWriter(path, engine='openpyxl', mode='w')

        self.dataFrame.to_excel(writer, sheet_name=self.name, index=False, startrow=2, header=False)

        # Costumize the excel a little
        redFill = PatternFill(start_color='FFF08080',
                              end_color='FFF08080',
                              fill_type='solid')

        blueFill = PatternFill(start_color='FF87CEFA',
                               end_color='FF87CEFA',
                               fill_type='solid')

        worksheet = writer.sheets[self.name]
        worksheet.merge_cells('B1:E1')
        worksheet['B1'] = "NON-COMMERCIAL"
        worksheet['B1'].alignment = Alignment(horizontal="center")
        worksheet.merge_cells('F1:I1')
        worksheet['F1'] = "Leveraged Money/Managed Money"
        worksheet['F1'].alignment = Alignment(horizontal="center")
        worksheet.merge_cells('J1:M1')
        worksheet['J1'] = "COMMERCIAL"
        worksheet['J1'].alignment = Alignment(horizontal="center")

        worksheet.merge_cells('B2:C2')
        worksheet['B2'] = "Long"
        worksheet['B2'].alignment = Alignment(horizontal="center")
        worksheet['B2'].fill = blueFill
        worksheet.merge_cells('D2:E2')
        worksheet['D2'] = "Short"
        worksheet['D2'].alignment = Alignment(horizontal="center")
        worksheet['D2'].fill = redFill

        worksheet.merge_cells('F2:G2')
        worksheet['F2'] = "Long"
        worksheet['F2'].alignment = Alignment(horizontal="center")
        worksheet['F2'].fill = blueFill
        worksheet.merge_cells('H2:I2')
        worksheet['H2'] = "Short"
        worksheet['H2'].alignment = Alignment(horizontal="center")
        worksheet['H2'].fill = redFill

        worksheet.merge_cells('J2:K2')
        worksheet['J2'] = "Long"
        worksheet['J2'].alignment = Alignment(horizontal="center")
        worksheet['J2'].fill = blueFill
        worksheet.merge_cells('L2:M2')
        worksheet['L2'] = "Short"
        worksheet['L2'].alignment = Alignment(horizontal="center")
        worksheet['L2'].fill = redFill

        #makes the cells a bit bigger
        for idx, col in enumerate(worksheet.columns, 1):
            worksheet.column_dimensions[get_column_letter(idx)].auto_size = True

        writer.close()