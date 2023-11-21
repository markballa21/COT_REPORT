from lxml import html
import requests
from datetime import datetime, timedelta
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter


class Web_COT_Data_Scrape():

    excel_file_name = "COT_REPORT"
    first_scrape = True

    def __init__(self, ticket, year):
        self.ticket = ticket
        self.year = year

        self.get_all_data()


    def web_scrape(self, date):

        url_legacy_futures = "https://www.tradingster.com/cot/legacy-futures/"+str(self.ticket)+"/"+str(date)
        response_legacy_futures = requests.get(url_legacy_futures)
        content_legacy_futures = response_legacy_futures.content
        parsed_content_legacy_futures = html.fromstring(content_legacy_futures)

        if(parsed_content_legacy_futures.xpath('/ html / head / title/text()')[0] == '500 Internal Server Error'):
            date = datetime.strptime(date, '%Y-%m-%d') + timedelta(days=-1)
            date = date.strftime("%Y-%m-%d")
            url_legacy_futures = "https://www.tradingster.com/cot/legacy-futures/" + str(self.ticket) + "/" + str(date)
            response_legacy_futures = requests.get(url_legacy_futures)
            content_legacy_futures = response_legacy_futures.content
            parsed_content_legacy_futures = html.fromstring(content_legacy_futures)
            if (parsed_content_legacy_futures.xpath('/ html / head / title/text()')[0] == '500 Internal Server Error'):
                return None

        Long_Non_Commercial = parsed_content_legacy_futures.xpath('/html/body/div[1]/div/div/div[3]/div[2]/table/tbody/tr[2]/td[1]/text()')[0].replace(',', '')
        Long_Non_Commercial_Change = \
            parsed_content_legacy_futures.xpath(
                '/html/body/div[1]/div/div/div[3]/div[2]/table/tbody/tr[4]/td[1]/span/text()')[0].replace(',', '')
        Long_Non_Commercial_Percent = parsed_content_legacy_futures.xpath('/html/body/div[1]/div/div/div[3]/div[2]/table/tbody/tr[6]/td[1]/text()')[0]
        Short_Non_Commercial = parsed_content_legacy_futures.xpath('/html/body/div[1]/div/div/div[3]/div[2]/table/tbody/tr[2]/td[2]/text()')[0].replace(',', '')
        Short_Non_Commercial_Change = \
        parsed_content_legacy_futures.xpath('/html/body/div[1]/div/div/div[3]/div[2]/table/tbody/tr[4]/td[2]/span/text()')[0].replace(',', '')
        Short_Non_Commercial_Percent = parsed_content_legacy_futures.xpath('/html/body/div[1]/div/div/div[3]/div[2]/table/tbody/tr[6]/td[2]/text()')[0]

        Long_Commercial = parsed_content_legacy_futures.xpath('/html/body/div[1]/div/div/div[3]/div[2]/table/tbody/tr[2]/td[4]/text()')[0].replace(',', '')
        Long_Commercial_Change = \
        parsed_content_legacy_futures.xpath('/html/body/div[1]/div/div/div[3]/div[2]/table/tbody/tr[4]/td[4]/span/text()')[0].replace(',', '')
        Long_Commercial_Percent = parsed_content_legacy_futures.xpath('/html/body/div[1]/div/div/div[3]/div[2]/table/tbody/tr[6]/td[4]/text()')[0]
        Short_Commercial = parsed_content_legacy_futures.xpath('/html/body/div[1]/div/div/div[3]/div[2]/table/tbody/tr[2]/td[5]/text()')[0].replace(',', '')
        Short_Commercial_Change = \
        parsed_content_legacy_futures.xpath('/html/body/div[1]/div/div/div[3]/div[2]/table/tbody/tr[4]/td[5]/span/text()')[0].replace(',', '')
        Short_Commercial_Percent = parsed_content_legacy_futures.xpath('/html/body/div[1]/div/div/div[3]/div[2]/table/tbody/tr[6]/td[5]/text()')[0]



        url_futures = "https://www.tradingster.com/cot/futures/fin/"+str(self.ticket)+"/"+str(date)
        response_futures = requests.get(url_futures)
        content_futures = response_futures.content
        parsed_content_futures = html.fromstring(content_futures)

        Long_Leveraged_Funds = parsed_content_futures.xpath('/html/body/div[1]/div/div/div[3]/div[2]/table/tbody/tr[3]/td[2]/text()')[0].replace(',', '')
        Long_Leveraged_Funds_Change = \
        parsed_content_futures.xpath('/html/body/div[1]/div/div/div[3]/div[2]/table/tbody/tr[3]/td[2]/span/text()')[0].replace(',', '')
        Long_Leveraged_Funds_Percent = parsed_content_futures.xpath('/html/body/div[1]/div/div/div[3]/div[2]/table/tbody/tr[3]/td[3]/text()')[0]
        Short_Leveraged_Funds = parsed_content_futures.xpath('/html/body/div[1]/div/div/div[3]/div[2]/table/tbody/tr[3]/td[5]/text()')[0].replace(',', '')
        Short_Leveraged_Change = \
        parsed_content_futures.xpath('/html/body/div[1]/div/div/div[3]/div[2]/table/tbody/tr[3]/td[5]/span/text()')[0].replace(',', '')
        Short_Leveraged_Funds_Percent = parsed_content_futures.xpath('/html/body/div[1]/div/div/div[3]/div[2]/table/tbody/tr[3]/td[6]/text()')[0]


        data = [date, int(Long_Non_Commercial)+int(Short_Non_Commercial), int(Long_Non_Commercial) - int(Short_Non_Commercial), 0,
                Long_Non_Commercial, Long_Non_Commercial_Change, Long_Non_Commercial_Percent, 0,
                Short_Non_Commercial, Short_Non_Commercial_Change, Short_Non_Commercial_Percent, 0,

                "LINE",

                int(Long_Leveraged_Funds) + int(Short_Leveraged_Funds), int(Long_Leveraged_Funds) - int(Short_Leveraged_Funds), 0,
                Long_Leveraged_Funds, Long_Leveraged_Funds_Change, Long_Leveraged_Funds_Percent, 0,
                Short_Leveraged_Funds, Short_Leveraged_Change, Short_Leveraged_Funds_Percent, 0,

                "LINE",

                int(Long_Commercial) + int(Short_Commercial), int(Long_Commercial) - int(Short_Commercial), 0,
                Long_Commercial, Long_Commercial_Change, Long_Commercial_Percent, 0,
                Short_Commercial, Short_Commercial_Change, Short_Commercial_Percent, 0,
                ]

        return data


    def get_all_data(self):
        list_of_dates = self.alltuesdays(self.year)

        #generate first row in excel
        df = pd.DataFrame(columns=np.arange(36))

        for date in list_of_dates:
            data = self.web_scrape(date)
            if(data == None):
                print("end data")
                break

            if(len(df) > 0):
                #
                last_data = df.loc[len(df) - 1]
                #
                data[3] = int(data[2]) - int(last_data[2])
                #
                data[7] = int(data[5]) + int(last_data[5])
                data[11] = int(data[9]) + int(last_data[9])

                data[15] = int(data[14]) - int(last_data[14])
                data[19] = int(data[17]) + int(last_data[17])
                data[23] = int(data[21]) + int(last_data[21])

                data[27] = int(data[26]) - int(last_data[26])
                data[31] = int(data[29]) + int(last_data[29])
                data[35] = int(data[33]) + int(last_data[33])

                print(data)

            df.loc[len(df)] = data

        df = df.loc[::-1]
        print(df)

        excel_file = openpyxl.Workbook()
        sheet = excel_file.active

        writer = pd.ExcelWriter('COT_REPORT.xlsx', engine='openpyxl')
        df.to_excel(writer, sheet_name='COT_REPORT', index=False, startrow=6, header=False)
        writer.close()


    def fill_excel(self):
        pass

    def load_excel(self):

        #Excel load
        excel_file = openpyxl.Workbook()
        sheet = excel_file.active
        excel_file['Sheet'].title = 'US30'

        sheet['A1'] = ""
        sheet['A2'] = ""
        sheet.freeze_panes = 'B1'

        # Start changing width from column C onwards
        column = 1
        while column < 26:
            i = get_column_letter(column)
            sheet.column_dimensions[i].width = 12
            column += 1

        #sheet.move_range("A7:Z7", rows=1, cols=0, translate=True)

        excel_file.save('COT_REPORT.xlsx')


    # DONT FORGET TO ADD IF THE SITE DIDNT FIND VALUES, TAKE A DAY BACK TO CHECK
    def alltuesdays(self, year):
        return pd.date_range(start=str(year), end=str(year + 1),
                freq='W-TUE').strftime('%Y-%m-%d').tolist()



